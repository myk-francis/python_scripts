#! python3
from pathlib import Path
from pydoc import cli
import re
import openpyxl
from openpyxl.styles import Font
import logging
from tqdm import tqdm
from datetime import datetime
import os

logging.basicConfig(filename='reportLOGS.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S %p %Z')

def add_alarms(rep_date, rep_time):

    pbar = tqdm(total=100, ascii=True, desc="Alarms")

    display_time = rep_time if (rep_time != '16') else 4

    display_am_pm = 'AM' if (rep_time != '16') else 'PM'

    file_path = f'Route Report ({datetime.now().strftime("%d-%m-%Y")}) {display_time} {display_am_pm}.xlsx'

    headings = ['No', 'Horse', 'Alarm type', 'Alarm time', 'Longitute', 'Latitude', 'Locate time', 'Speed', 'Location']

    alarms_data = [headings]

    pbar.update(20)


    try:
        if os.path.exists(str(Path.cwd()) + '/INPUT/alarms.xlsx'):
            pass
        else:
            print("\n\nAlarms file missing from INPUT folder..\nMake sure extension is XLSX(xlsx not xls)!!!\n\n")
            return False

        pbar.update(20)

        alarms_wb = openpyxl.load_workbook('INPUT/alarms.xlsx')

        index_alarms = 0

        for sheet in alarms_wb.sheetnames:

            active_sheet = alarms_wb[sheet]

            for row in range(2, active_sheet.max_row + 1):

                row_data = []

                for col in range(1, active_sheet.max_column + 1):
                    
                    if col >= 2 and col < 9:
                        row_data.append(active_sheet.cell(row=row, column=col).value)
                            
                
                if all(i is None for i in row_data):
                    break
                
                index_alarms += 1
                row_data.insert(0, index_alarms)
                row_data.insert(1, active_sheet.title)
                alarms_data.append(row_data)

        

        pbar.update(20)

        route_rep_wb = openpyxl.load_workbook(f'Route Report ({datetime.now().strftime("%d-%m-%Y")}) {display_time} {display_am_pm}.xlsx')

        route_rep_wb.create_sheet(index=2, title='Alarms')

        alarms_sheet = route_rep_wb['Alarms']

        row = alarms_sheet.row_dimensions[1]

        row.font = Font(size=15, bold=True)

        pbar.update(20)

        for row in alarms_data:
            alarms_sheet.append(row)

        for row in range(2, alarms_sheet.max_row + 1):
            latitude = alarms_sheet['F' + str(row)].value
            longitude = alarms_sheet['E' + str(row)].value
            alarms_sheet['J' + str(row)].hyperlink = f"http://maps.google.com/maps?q={latitude},{longitude}"

        route_rep_wb.save(f'Route Report ({datetime.now().strftime("%d-%m-%Y")}) {display_time} {display_am_pm}.xlsx')

        pbar.update(20)
        pbar.close()

    except Exception as e:
        print(e)
        logging.debug('--------------------ERROR---------------------')
        logging.debug(e)
        print("\n\nSomething went wrong!!!")
        return False    
    else:
        if os.path.exists(str(Path.cwd()) + '/INPUT/alarms.xlsx'):
            os.remove(str(Path.cwd()) + '/INPUT/alarms.xlsx')
        return True
