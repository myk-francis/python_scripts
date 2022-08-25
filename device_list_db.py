import psycopg2
import openpyxl
from pathlib import Path

def upload_device_list():

    try:
        connection = psycopg2.connect(user="postgres",
                                    password="myk",
                                    host="127.0.0.1",
                                    port="5432",
                                    database="test_db")
        cursor = connection.cursor()
        print("Connection to DATABASE was established!\n")
        print("Reading Manifest and Device List ...\n\n")

        #creating a cursor
        cursor = connection.cursor()

        device_data = []
        updated_list = []
        
        # device_wb = openpyxl.load_workbook(f'{str(Path.cwd())}/INPUT/device_list.xlsx')
        # device_sheet = device_wb.active

        device_wb = openpyxl.load_workbook(f'{str(Path.cwd())}/INPUT/device_list_update.xlsx')
        device_sheet = device_wb.active


        # for row in range(2, 2000):

        #     row_data = []

        #     for col in range(1, 6):
        #         row_data.append(device_sheet.cell(row=row, column=col).value)
                        
        #     if all(i is None for i in row_data):
        #         break

        #     row_data.pop(3)

        #     for j in range(len(row_data)):
        #         if (row_data[j] == None):
        #             row_data[j] = ''
                    
        #     device_data.append(tuple(row_data)))

        for row in range(2, 2000):

            row_data = []

            for col in range(1, 4):
                row_data.append(device_sheet.cell(row=row, column=col).value)
                        
            if all(i is None for i in row_data):
                break

            for j in range(len(row_data)):
                if (row_data[j] == None):
                    row_data[j] = ''
                    
            device_data.append(row_data)


        
        #The first insert into the table 
        # args = ','.join(cursor.mogrify("(%s,%s,%s,%s,%s)", i).decode('utf-8')
        #                 for i in device_data)
        
        # executing the sql statement
        # cursor.execute("INSERT INTO device_list (device_id , sim_card, iccid, current_account, horse) VALUES " + (args))


        # Update single record now device_list
        sql_select_query = """SELECT * FROM device_list ORDER BY id ASC"""
        cursor.execute(sql_select_query)
        records = cursor.fetchall()

        for record in records:
            for excel_data in device_data:
                if (str(record[1]) == str(excel_data[0])) and (str(record[5]) != str(excel_data[1])):
                    sql_update_query = """UPDATE device_list SET 
                                            horse = %s
                                        WHERE device_id = %s"""
                    cursor.execute(sql_update_query, (excel_data[1],record[1]))

                if (str(record[1]) == str(excel_data[0])) and (str(record[4]) != str(excel_data[2])):
                    sql_update_query = """UPDATE device_list SET 
                                            current_account = %s
                                        WHERE device_id = %s"""
                    cursor.execute(sql_update_query, (excel_data[2],record[1]))
        
        # commiting changes
        connection.commit()

        print("Record updated successfully into device_list table...\n\n")

    except (Exception, psycopg2.Error) as error:
        print("Failed to update record into device_list table\n\n", error)

    finally:
        # closing database connection.
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed\n\n")