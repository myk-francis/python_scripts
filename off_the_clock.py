#! python3

import glob
import openpyxl
import logging
from pathlib import Path
import shutil, os

import time

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S %p %Z')

user_input = input('(a)ll (o)ff_the_clock or (q)uit:')
if user_input == 'q':
    exit()


if user_input == 'o':
    logging.debug("Generating Off the clock alarms")
else:
    logging.debug("Generating for all alarms")
    

merged_wb = openpyxl.load_workbook("alarms.xlsx")

merged_sheet = merged_wb.active

alarms_data = []
merged_alarms_list = []

for sheet in merged_wb.sheetnames:

    active_sheet = merged_wb[sheet]

    for row in range(2, active_sheet.max_row + 1):

        row_data = []

        for col in range(1, active_sheet.max_column + 1):
            
            if col >= 2 and col < 9:
                row_data.append(active_sheet.cell(row=row, column=col).value)
                    
        
        if all(i is None for i in row_data):
            break
        
        row_data.insert(0, active_sheet.title)
        alarms_data.append(row_data)



for alarms in alarms_data:
    latitude = alarms[4]
    longitude = alarms[3]
    alarms.pop(5)

    merged_alarms_list.append(alarms + [f"http://maps.google.com/maps?q={latitude},{longitude}"])

merged_wb = openpyxl.Workbook()

merged_sheet = merged_wb.active

for row in merged_alarms_list:
    merged_sheet.append(row)

merged_wb.save('merged_alarms.xlsx')

if os.path.exists('alarms.xlsx'):
    os.remove('alarms.xlsx')

time.sleep(3)

wb = openpyxl.load_workbook("merged_alarms.xlsx")
sheet = wb.active

output_file = open('output.txt', 'w')


for row in range(1, sheet.max_row + 1):
    off_the_clock_list = []
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value == None:
            break

        if user_input == 'a':
            if col == 1:
                output_file.write(f'Truck: {sheet.cell(row=row, column=col).value}\n')
            if col == 2:
                output_file.write(f'Alarm Type: {sheet.cell(row=row, column=col).value}\n')
            if col == 3:
                output_file.write(f'Alarm Time: {sheet.cell(row=row, column=col).value}\n')
            if col == 6:
                output_file.write(f'Speed: {sheet.cell(row=row, column=col).value}\n')
            if col == 7:
                output_file.write(f'Location: {sheet.cell(row=row, column=col).value}\n')
            if col == 8:
                output_file.write(f'Link: {sheet.cell(row=row, column=col).value}\n')

        if user_input == 'o':
            if sheet.cell(row=row, column=2).value == 'Off the clock':
                
                if float(sheet.cell(row=row, column=6).value) < 10:
                    break

                if col == 1:
                    output_file.write(f'Truck: {sheet.cell(row=row, column=col).value}\n')
                if col == 2:
                    output_file.write(f'Alarm Type: {sheet.cell(row=row, column=col).value}\n')
                if col == 3:
                    output_file.write(f'Alarm Time: {sheet.cell(row=row, column=col).value}\n')
                if col == 6:
                    output_file.write(f'Speed: {sheet.cell(row=row, column=col).value}\n')
                if col == 7:
                    output_file.write(f'Location: {sheet.cell(row=row, column=col).value}\n')
                if col == 8:
                    output_file.write(f'Link: {sheet.cell(row=row, column=col).value}\n')

    output_file.write('\n\n')

        

output_file.close()
logging.debug("Check the output file")

user_input = input("\n\nIs everything GUCCI??? If Yes I'll delete the alarms.xlsx file?\n\n1.Yes\n2.No\n>")
if user_input == 1:
    if os.path.exists('merged_alarms.xlsx'):
        os.remove('merged_alarms.xlsx')
    