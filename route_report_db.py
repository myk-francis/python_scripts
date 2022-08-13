import psycopg2
import openpyxl
from pathlib import Path

def upload_report():
        
    try:
        connection = psycopg2.connect(user="postgres",
                                    password="myk",
                                    host="127.0.0.1",
                                    port="5432",
                                    database="test_db")
        cursor = connection.cursor()
        print("Connection to DATABASE was established!\n")
        print("Reading Route Report...\n\n")

        #creating a cursor
        cursor = connection.cursor()

        route_data = []

        
        client_wb = openpyxl.load_workbook(f'{str(Path.cwd())}/INPUT/client_report.xlsx')
        client_sheet = client_wb.active

        for row in range(3, 1000):

            row_data = []

            for col in range(1, client_sheet.max_column + 1):
                row_data.append(client_sheet.cell(row=row, column=col).value)
                        
            if all(i is None for i in row_data):
                break

            row_data.pop(0)
            
            for j in range(len(row_data)):
                if (row_data[j] == None) and (j != 5):
                    row_data[j] = ''
                    
            route_data.append(tuple(row_data))
        
        # list of rows to be inserted
        # cursor.mogrify() to insert multiple values
        args = ','.join(cursor.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)", i).decode('utf-8')
                        for i in route_data)
        
        # executing the sql statement
        cursor.execute("INSERT INTO route_tracking (horse , device_id , battery , device_status , device_location , location_time , speed , geo_fence , date_from , date_to , distance) VALUES " + (args))
        
        # commiting changes
        connection.commit()

        print("Record inserted successfully into route table...\n\n")

    except (Exception, psycopg2.Error) as error:
        print("Failed to insert record into route table\n\n", error)

    finally:
        # closing database connection.
        if connection:
            cursor.close()
            connection.close()
            print("PostgreSQL connection is closed\n\n")