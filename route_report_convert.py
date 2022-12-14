#! python3
from pathlib import Path
from pydoc import cli
import re
import openpyxl
from openpyxl.styles import Font
import logging
from tqdm import tqdm
from datetime import datetime
import os
    
logging.basicConfig(filename='reportLOGS.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%Y-%m-%d %H:%M:%S %p %Z')

def convert_report(rep_date, rep_time):

    try:
        print("\n")
        pbar = tqdm(total=100, ascii=True, desc="Converting")

        display_time = rep_time if (rep_time != '16') else 4

        display_am_pm = 'AM' if (rep_time != '16') else 'PM'


        manifest_data = []
        client_data = []
        tz_compiled_data = []
        sa_compiled_data = []
        za_compiled_data = []
        missing_trucks = []
        missing_trucks_locator = []
        heading = ['NO', 'HORSE', 'TRAILER #1', 'TRAILER #2', 'CARGO TYPE', 'DESTINATION', 'TRANSPORTER', 'Device ID','Battery', 'Status','Location','Location Time', 'Speed', 'Geo Fence', 'From', 'To','KM']


        manifest_wb = openpyxl.load_workbook(f'{str(Path.cwd())}/INPUT/manifest.xlsx')
        sheet = manifest_wb.active

        client_wb = openpyxl.load_workbook(f'{str(Path.cwd())}/INPUT/client_report.xlsx')
        client_sheet = client_wb.active

        pbar.update(10)

        for row in range(2, 1000):

            row_data = []

            for col in range(1, sheet.max_column + 1):
                
                if col > 2 and col < 11:
                    row_data.append(sheet.cell(row=row, column=col).value)
                        
            
            if all(i is None for i in row_data):
                break
            manifest_data.append(row_data)

        pbar.update(10)
        # print(manifest_data) -- FOR DEBUGGING

        for row in range(3, 1000):

            row_data = []

            for col in range(1, client_sheet.max_column + 1):
                row_data.append(client_sheet.cell(row=row, column=col).value)
            
            if all(i is None for i in row_data):
                break
            client_data.append(row_data)

        pbar.update(10)

        #Ordering the list
        myorder = [5, 6, 7, 0, 1, 2, 4, 3]
        duplicate_manifest_data = []
        for data in manifest_data:
            data = [data[i] for i in myorder]
            duplicate_manifest_data.append(data)

        pbar.update(10)

        manifest_data = duplicate_manifest_data

        DRR = '(DRR)'
        TCR = '(TCR)'
        index_t = 0
        index_s = 0
        index_z = 0
        for client in client_data:
            if DRR in client[1]:
                client[1] = re.sub(' \(DRR\)', '', client[1])
            
            if TCR in client[1]:
                client[1] = re.sub(' \(TCR\)', '', client[1])

            missing_trucks_locator = []
            for manifest in manifest_data:
                if client[1] in manifest:
                    # print(client)
                    # print(manifest)
                    if manifest[7] == 'TANZANIA':
                        index_t += 1
                        manifest.pop(7)
                        manifest.pop(4)
                        sliced_client = client[2:12]
                        tz_compiled_data.append([index_t] + manifest + sliced_client)
                    elif manifest[7] == 'SOUTH AFRICA':
                        index_s += 1
                        manifest.pop(7)
                        manifest.pop(4)
                        sliced_client = client[2:12]
                        sa_compiled_data.append([index_s] + manifest + sliced_client)
                    elif manifest[7] == 'ZAMBIA':
                        index_z += 1
                        manifest.pop(7)
                        manifest.pop(4)
                        sliced_client = client[2:12]
                        za_compiled_data.append([index_z] + manifest + sliced_client)

                    missing_trucks_locator.append(True)
                else:
                    missing_trucks_locator.append(False)
            
            if any(missing_trucks_locator) == False:
                if client[1] not in missing_trucks:
                    missing_trucks.append(client[1:3])

        pbar.update(10)

        wb = openpyxl.Workbook()

        dest_filename = f'Route Report ({datetime.now().strftime("%d-%m-%Y")}) {display_time} {display_am_pm}.xlsx'

        # dest_filename = f'route_report_output.xlsx'
        pbar.update(10)

        sheet = wb.active

        sheet.title = 'DRC - TZ'

        tz_compiled_data.insert(0, heading)

        pbar.update(10)

        for row in tz_compiled_data:
            sheet.append(row)

        row = sheet.row_dimensions[1]

        row.font = Font(size=15, bold=True)

        sheet2 = wb.create_sheet(title="DRC - SA")

        sa_compiled_data.insert(0, heading)

        for row in sa_compiled_data:
            sheet2.append(row)

        row = sheet2.row_dimensions[1]

        row.font = Font(size=15, bold=True)

        sheet3 = wb.create_sheet(title="DRC - ZA")

        za_compiled_data.insert(0, heading)

        for row in za_compiled_data:
            sheet3.append(row)

        row = sheet3.row_dimensions[1]

        row.font = Font(size=15, bold=True)

        wb.save(filename = dest_filename)

        pbar.update(10)

        pbar.update(10)

        if len(missing_trucks) > 0:
            print("ERROR --> MISSING TRUCKS")
            print(f'MISSING TRUCKS TOTAL({len(missing_trucks)}) \n\n\n {missing_trucks} \n\n\n')
            logging.debug(f'MISSING TRUCKS TOTAL({len(missing_trucks)}) \n\n\n {missing_trucks} \n\n\n')
        else:
            if os.path.exists(str(Path.cwd()) + '/INPUT/client_report.xlsx'):
                os.remove(str(Path.cwd()) + '/INPUT/client_report.xlsx')

            logging.debug("SUCCESSFULL")
            logging.debug("Please check file route_report_output.xlsx for the output")
            logging.debug("------END OF PROGRAM-------")

        pbar.update(10)
        pbar.close()
        print("\n")
    except Exception as e:
        print(e)
        return False
    else:
        return True