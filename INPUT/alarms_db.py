import os
import psycopg2
import openpyxl
from pathlib import Path

try:
    connection = psycopg2.connect(user="postgres",
                                  password="myk",
                                  host="127.0.0.1",
                                  port="5432",
                                  database="test_db")
    cursor = connection.cursor()
    print("Connection to DATABASE was established!\n")
    print("Reading Alarms...\n\n")

    if os.path.exists(str(Path.cwd()) + '/INPUT/alarms.xlsx'):
        pass
    else:
        print("\n\nAlarms file missing from INPUT folder..\nMake sure extension is XLSX(xlsx not xls)!!!\n\n")
        exit()

    #creating a cursor
    cursor = connection.cursor()

    alarms_wb = openpyxl.load_workbook('INPUT/alarms.xlsx')

    alarms_data = []

    for sheet in alarms_wb.sheetnames:

        active_sheet = alarms_wb[sheet]

        for row in range(2, active_sheet.max_row + 1):

            row_data = []

            for col in range(1, active_sheet.max_column + 1):
                
                if col >= 2 and col < 9:
                    row_data.append(active_sheet.cell(row=row, column=col).value)
                        
            
            if all(i is None for i in row_data):
                break
            
            row_data.insert(0, active_sheet.title)
            alarms_data.append(row_data)

    
    for i in range(len(alarms_data)):
        for j in range(len(alarms_data[i])):
            if alarms_data[i][j] == "Into the community alarm":
                alarms_data[i][j] = "Enter Fence alarm"

            if alarms_data[i][j] == "Regional alarm":
                alarms_data[i][j] = "Outgoing Fence alarm"
    
    # list of rows to be inserted
    # cursor.mogrify() to insert multiple values
    args = ','.join(cursor.mogrify("(%s,%s,%s,%s,%s,%s,%s,%s)", i).decode('utf-8')
                    for i in alarms_data)
    
    # executing the sql statement
    cursor.execute("INSERT INTO alarms (horse , alarm_type, alarm_time , longitude , latitude , locate_time , speed , device_location) VALUES " + (args))
    
    # commiting changes
    connection.commit()

    print("Record inserted successfully into alarms table...\n\n")

except (Exception, psycopg2.Error) as error:
    print("Failed to insert record into alarms table\n\n", error)

finally:
    # closing database connection.
    if connection:
        cursor.close()
        connection.close()
        print("PostgreSQL connection is closed\n\n")